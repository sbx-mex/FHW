#!/usr/bin/env python3
"""Construye el modelo auditable de FHW · Cada Taza Cuenta.

Calcula FHW / Bebidas Lobby por tienda y promedia los porcentajes válidos
del alcance seleccionado.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import shutil
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "input"
OUTPUT = ROOT / "public" / "data"
TARGET = 0.10
MIN_SYNC_RATE = 0.90
MONTHS = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean(value))
    return "".join(char for char in text.casefold() if not unicodedata.combining(char))


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        result = float(value)
    else:
        text = clean(value).replace(",", "").replace("%", "")
        try:
            result = float(text)
        except ValueError:
            return None
    return result if math.isfinite(result) else None


def percent(value: Any) -> float | None:
    result = number(value)
    if result is None:
        return None
    if isinstance(value, str) and "%" in value:
        return result / 100
    return result / 100 if abs(result) > 1.5 else result


def ceco(value: Any) -> str | None:
    parsed = number(value)
    if parsed is not None:
        return str(int(round(parsed)))
    digits = re.sub(r"\D", "", clean(value))
    return digits or None


def parse_period(value: Any) -> tuple[int, int] | None:
    parsed = number(value)
    if parsed is None:
        return None
    raw = int(parsed)
    if raw > 1000:
        return raw // 100, raw % 100
    return 2026, raw


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def atomic_json(path: Path, payload: Any, *, compact: bool = True) -> None:
    """Publica JSON completo o conserva la versión anterior si el proceso falla."""
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(f"{path.suffix}.next")
    text = json.dumps(payload, ensure_ascii=False, separators=(",", ":")) if compact else json.dumps(payload, ensure_ascii=False, indent=2)
    temporary.write_text(text, encoding="utf-8")
    temporary.replace(path)


def decoded_lines(path: Path) -> list[str]:
    head = path.read_bytes()[:4]
    encoding = "utf-16" if head.startswith((b"\xff\xfe", b"\xfe\xff")) else "utf-8-sig"
    return path.read_text(encoding=encoding).splitlines()


def report_rows(path: Path, required: Iterable[str]) -> tuple[list[dict[str, str]], int]:
    lines = decoded_lines(path)
    required_keys = {key(item) for item in required}
    header_index = None
    for index, line in enumerate(lines):
        parsed = next(csv.reader([line]))
        if required_keys.issubset({key(item) for item in parsed}):
            header_index = index
            break
    if header_index is None:
        raise ValueError(f"{path.name}: no se encontró el encabezado requerido")
    rows = list(csv.DictReader(lines[header_index:]))
    return rows, header_index + 1


def value_from(row: dict[str, Any], *names: str) -> Any:
    mapped = {key(name): value for name, value in row.items()}
    for name in names:
        if key(name) in mapped:
            return mapped[key(name)]
    raise KeyError(f"No se encontró columna: {', '.join(names)}")


def aggregate_report(path: Path, measure: tuple[str, ...]) -> tuple[dict[tuple[str, int], float], dict[str, Any]]:
    rows, header_row = report_rows(path, ("Tiendas", "Semana", measure[0]))
    grouped: defaultdict[tuple[str, int], float] = defaultdict(float)
    source_rows = duplicate_rows = invalid_rows = 0
    for row in rows:
        source_rows += 1
        store = ceco(value_from(row, "Tiendas", "Tienda", "CeCo"))
        period = parse_period(value_from(row, "Semana"))
        amount = number(value_from(row, *measure))
        if not store or not period or amount is None:
            invalid_rows += 1
            continue
        _, week = period
        if (store, week) in grouped:
            duplicate_rows += 1
        grouped[(store, week)] += amount
    return dict(grouped), {
        "file": path.name,
        "headerRow": header_row,
        "rows": source_rows,
        "validKeys": len(grouped),
        "duplicatesConsolidated": duplicate_rows,
        "invalidRows": invalid_rows,
        "sha256": sha256(path),
    }


def locate_header(sheet: Any, required: Iterable[str], max_rows: int = 10) -> tuple[int, dict[str, int]]:
    required_keys = {key(item) for item in required}
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True), 1):
        headers = {key(value): index for index, value in enumerate(row) if clean(value)}
        if required_keys.issubset(headers):
            return row_number, headers
    raise ValueError(f"{sheet.title}: encabezado no encontrado")


def column(headers: dict[str, int], *names: str, required: bool = True) -> int | None:
    for name in names:
        if key(name) in headers:
            return headers[key(name)]
    if required:
        raise ValueError(f"Falta columna: {', '.join(names)}")
    return None


def is_yes(value: Any) -> bool:
    return key(value) in {"si", "sí", "1", "true", "aplica", "yes"}


def load_reference(path: Path) -> tuple[dict[str, dict[str, str]], dict[int, str], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        directory_sheet = workbook["Directorio"] if "Directorio" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        # El directorio operativo puede nombrar la llave como CC o CeCo.
        header_row, headers = locate_header(directory_sheet, ("Tienda", "DM"))
        cc_i = column(headers, "CC", "CeCo")
        store_i = column(headers, "Tienda", "Nombre de Tienda")
        region_i = column(headers, "Región", "Region")
        dm_i = column(headers, "DM")
        status_i = column(headers, "Estatus", required=False)
        applies_i = column(headers, "Aplica", "Aplica FHW", "Aplica Si", required=False)
        directory: dict[str, dict[str, str]] = {}
        excluded = 0
        for row in directory_sheet.iter_rows(min_row=header_row + 1, values_only=True):
            code = ceco(row[cc_i])
            if not code or code in directory:
                continue
            applies = is_yes(row[applies_i]) if applies_i is not None else (
                key(row[status_i]) in {"abierta", "activo", "activa"} if status_i is not None else True
            )
            if not applies:
                excluded += 1
                continue
            directory[code] = {
                "store": clean(row[store_i]) or f"Tienda {code}",
                "region": clean(row[region_i]) or "Sin región",
                "dm": clean(row[dm_i]) or "Sin DM",
            }

        calendar_sheet = None
        for name in ("Base_Año_Mes_Sem", "Base_Mes_Semana", "Base Año Mes Semana"):
            if name in workbook.sheetnames:
                calendar_sheet = workbook[name]
                break
        week_month: dict[int, str] = {}
        if calendar_sheet is not None:
            cal_header, cal_headers = locate_header(calendar_sheet, ("Año", "Mes", "Semana"))
            year_i = column(cal_headers, "Año")
            month_i = column(cal_headers, "Mes")
            week_i = column(cal_headers, "Semana")
            for row in calendar_sheet.iter_rows(min_row=cal_header + 1, values_only=True):
                year = number(row[year_i])
                week = number(row[week_i])
                month = clean(row[month_i]).title()[:3]
                if year == 2026 and week is not None and month in MONTHS:
                    week_month[int(week)] = month

        historical: list[dict[str, Any]] = []
        if "CTC_Tienda" in workbook.sheetnames:
            sheet = workbook["CTC_Tienda"]
            hist_header, hist_headers = locate_header(sheet, ("Año", "Semana", "Ceco", "Part FHW"))
            year_i = column(hist_headers, "Año")
            week_i = column(hist_headers, "Semana")
            ceco_i = column(hist_headers, "Ceco", "CeCo")
            ratio_i = column(hist_headers, "Part FHW", "FHW")
            for row in sheet.iter_rows(min_row=hist_header + 1, values_only=True):
                year, week, code, ratio = number(row[year_i]), number(row[week_i]), ceco(row[ceco_i]), percent(row[ratio_i])
                if year == 2026 and week is not None and 1 <= int(week) <= 29 and code in directory and ratio is not None and 0 <= ratio <= 1:
                    historical.append({"year": 2026, "week": int(week), "ceco": code, "ratio": ratio})

        historical_dm: list[dict[str, Any]] = []
        if "CTC_DM" in workbook.sheetnames:
            sheet = workbook["CTC_DM"]
            dm_header, dm_headers = locate_header(sheet, ("Año", "Semana", "DM", "Part FHW"))
            year_i = column(dm_headers, "Año")
            week_i = column(dm_headers, "Semana")
            dm_i = column(dm_headers, "DM")
            ratio_i = column(dm_headers, "Part FHW", "FHW")
            valid_dms = {item["dm"] for item in directory.values()}
            for row in sheet.iter_rows(min_row=dm_header + 1, values_only=True):
                year, week, dm, ratio = number(row[year_i]), number(row[week_i]), clean(row[dm_i]), percent(row[ratio_i])
                if year == 2026 and week is not None and 1 <= int(week) <= 29 and dm in valid_dms and ratio is not None and 0 <= ratio <= 1:
                    historical_dm.append({"year": 2026, "week": int(week), "dm": dm, "ratio": ratio})

        return directory, week_month, historical, historical_dm, {
            "file": path.name,
            "sheet": directory_sheet.title,
            "includedStores": len(directory),
            "excludedStores": excluded,
            "applyRule": "Aplica = Sí" if applies_i is not None else "Estatus = Abierta (equivale a Aplica = Sí)",
            "applyColumnPresent": applies_i is not None,
            "calendarWeeks": len(week_month),
            "historicalStoreRows": len(historical),
            "historicalDmRows": len(historical_dm),
            "sha256": sha256(path),
        }
    finally:
        workbook.close()


def load_uploaded_historical(path: Path, directory: dict[str, dict[str, str]]) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows, _ = report_rows(path, ("Año", "Semana", "Ceco", "FHW"))
    result = []
    for row in rows:
        year = number(value_from(row, "Año"))
        week = number(value_from(row, "Semana"))
        code = ceco(value_from(row, "Ceco", "CeCo"))
        ratio = percent(value_from(row, "FHW"))
        if year == 2026 and week is not None and 1 <= int(week) <= 29 and code in directory and ratio is not None and 0 <= ratio <= 1:
            result.append({"year": 2026, "week": int(week), "ceco": code, "ratio": ratio})
    return result


def load_calendar_override(path: Path) -> dict[int, str]:
    if not path.exists():
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = next((workbook[name] for name in ("Base_Año_Mes_Sem", "Base_Mes_Semana", "Base Año Mes Semana") if name in workbook.sheetnames), workbook[workbook.sheetnames[0]])
        header_row, headers = locate_header(sheet, ("Año", "Mes", "Semana"))
        year_i, month_i, week_i = column(headers, "Año"), column(headers, "Mes"), column(headers, "Semana")
        result = {}
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            year, week, month = number(row[year_i]), number(row[week_i]), clean(row[month_i]).title()[:3]
            if year == 2026 and week is not None and month in MONTHS:
                result[int(week)] = month
        return result
    finally:
        workbook.close()


def month_fallback(week: int) -> str:
    # ISO-week midpoint; used only when the supplied calendar has no mapping.
    return datetime.fromisocalendar(2026, week, 4).strftime("%b").title().replace("Jan", "Ene").replace("Apr", "Abr").replace("Aug", "Ago").replace("Dec", "Dic")


def average_ratio(items: Iterable[dict[str, Any]]) -> float:
    values = [float(item["ratio"]) for item in items if 0 <= float(item["ratio"]) <= 1]
    return sum(values) / len(values) if values else 0


def build() -> dict[str, Any]:
    fhw_path = INPUT / "CTC_FHW.csv"
    lobby_path = INPUT / "CTC_Bebidas_Lobby.csv"
    reference_path = INPUT / "Directorio_FHW.xlsx"
    for path in (fhw_path, lobby_path, reference_path):
        if not path.exists():
            raise FileNotFoundError(f"Falta fuente obligatoria: {path.relative_to(ROOT)}")

    fhw, fhw_audit = aggregate_report(fhw_path, ("Cantidad Ajustada", "FHW"))
    lobby, lobby_audit = aggregate_report(lobby_path, ("Unidades", "Bebidas Lobby"))
    directory, week_month, historical, historical_dm, reference_audit = load_reference(reference_path)
    calendar_override = load_calendar_override(INPUT / "Base_Año_Mes_Sem.xlsx")
    if calendar_override:
        week_month = calendar_override
        reference_audit["calendarWeeks"] = len(week_month)
        reference_audit["calendarSource"] = "Base_Año_Mes_Sem.xlsx"
    else:
        reference_audit["calendarSource"] = reference_path.name
    uploaded_historical = load_uploaded_historical(INPUT / "FHW_Sem1_29.csv", directory)
    if uploaded_historical:
        historical = uploaded_historical

    hierarchy_map: defaultdict[str, defaultdict[str, list[dict[str, str]]]] = defaultdict(lambda: defaultdict(list))
    for code, item in sorted(directory.items(), key=lambda pair: (pair[1]["region"].casefold(), pair[1]["dm"].casefold(), pair[1]["store"].casefold())):
        hierarchy_map[item["region"]][item["dm"]].append({"ceco": code, "name": item["store"], "applies": True})
    hierarchy = [
        {
            "name": region,
            "dms": [{"name": dm, "stores": stores} for dm, stores in dms.items()],
        }
        for region, dms in hierarchy_map.items()
    ]

    source_weeks = sorted({week for _, week in set(fhw) | set(lobby) if week >= 30})
    synchronization = []
    ready_weeks = []
    for item_week in source_weeks:
        fhw_codes = {code for code, week in fhw if week == item_week}
        lobby_codes = {code for code, week in lobby if week == item_week}
        matched_codes = fhw_codes & lobby_codes
        sync_rate = len(matched_codes) / max(len(fhw_codes), len(lobby_codes), 1)
        has_both_sources = bool(fhw_codes) and bool(lobby_codes)
        status = "ready" if has_both_sources and sync_rate >= MIN_SYNC_RATE else "pending"
        if has_both_sources and sync_rate < MIN_SYNC_RATE:
            raise ValueError(
                f"Semana {item_week} fuera de congruencia: "
                f"FHW={len(fhw_codes)}, Lobby={len(lobby_codes)}, coincidencias={len(matched_codes)}"
            )
        if status == "ready":
            ready_weeks.append(item_week)
        synchronization.append({
            "week": item_week,
            "status": status,
            "fhwStores": len(fhw_codes),
            "lobbyStores": len(lobby_codes),
            "matchedStores": len(matched_codes),
            "onlyFhw": len(fhw_codes - lobby_codes),
            "onlyLobby": len(lobby_codes - fhw_codes),
            "matchRate": round(sync_rate, 8),
        })

    live_keys = sorted(
        ((set(fhw) & set(lobby)) & {(code, week) for code, week in set(fhw) | set(lobby) if week in ready_weeks}),
        key=lambda item: (item[1], int(item[0])),
    )
    live_records: list[dict[str, Any]] = []
    excluded_no_directory = zero_denominator = 0
    for code, week in live_keys:
        if week < 30:
            continue
        meta = directory.get(code)
        if meta is None:
            excluded_no_directory += 1
            continue
        denominator = lobby[(code, week)]
        if denominator <= 0:
            zero_denominator += 1
            continue
        numerator = fhw[(code, week)]
        live_records.append({
            "year": 2026,
            "week": week,
            "month": week_month.get(week, month_fallback(week)),
            "ceco": code,
            "store": meta["store"],
            "dm": meta["dm"],
            "region": meta["region"],
            "fhw": round(numerator, 6),
            "lobby": round(denominator, 6),
            "ratio": round(numerator / denominator, 8),
            "source": "calculado",
        })

    historical_records = []
    for item in historical:
        meta = directory[item["ceco"]]
        historical_records.append({
            **item,
            "month": week_month.get(item["week"], month_fallback(item["week"])),
            "store": meta["store"],
            "dm": meta["dm"],
            "region": meta["region"],
            "fhw": None,
            "lobby": None,
            "ratio": round(item["ratio"], 8),
            "source": "histórico directo",
        })

    all_records = sorted(historical_records + live_records, key=lambda item: (item["week"], item["store"].casefold()))

    # La serie completa es comparable porque parte del porcentaje por tienda.
    average_rollups: dict[str, list[dict[str, Any]]] = {"national": [], "region": [], "dm": []}
    for level in ("national", "region", "dm"):
        grouped_adoption: defaultdict[tuple[int, str], list[dict[str, Any]]] = defaultdict(list)
        for item in all_records:
            name = "Nacional" if level == "national" else item[level]
            grouped_adoption[(item["week"], name)].append(item)
        for (item_week, name), items in sorted(grouped_adoption.items()):
            stores = {item["ceco"] for item in items}
            above = sum(1 for item in items if item["ratio"] > TARGET)
            average_rollups[level].append({
                "week": item_week,
                "name": name,
                "stores": len(stores),
                "aboveTarget": above,
                "ratio": round(average_ratio(items), 8),
            })
    live_weeks = sorted({item["week"] for item in live_records})
    latest_week = max(live_weeks) if live_weeks else 0
    latest = [item for item in live_records if item["week"] == latest_week]
    total_fhw = sum(item["fhw"] for item in latest)
    total_lobby = sum(item["lobby"] for item in latest)
    average_latest = average_ratio(latest)
    coverage_by_week = []
    executive_weeks = []
    for item_week in live_weeks:
        fhw_keys = {code for code, source_week in fhw if source_week == item_week}
        lobby_keys = {code for code, source_week in lobby if source_week == item_week}
        published = [item for item in live_records if item["week"] == item_week]
        coverage_by_week.append({
            "week": item_week,
            "fhwStores": len(fhw_keys),
            "lobbyStores": len(lobby_keys),
            "matchedStores": len(fhw_keys & lobby_keys),
            "publishedStores": len(published),
        })
        week_fhw = sum(item["fhw"] for item in published)
        week_lobby = sum(item["lobby"] for item in published)
        executive_weeks.append({
            "week": item_week,
            "month": week_month.get(item_week, month_fallback(item_week)),
            "fhw": round(week_fhw, 6),
            "lobby": round(week_lobby, 6),
            "ratio": round(average_ratio(published), 8),
            "stores": len(published),
            "aboveTarget": sum(1 for item in published if item["ratio"] > TARGET),
            "nearTarget": sum(1 for item in published if 0.08 <= item["ratio"] <= TARGET),
            "opportunity": sum(1 for item in published if item["ratio"] < 0.08),
        })

    source_duplicates = fhw_audit["duplicatesConsolidated"] + lobby_audit["duplicatesConsolidated"]
    source_invalid = fhw_audit["invalidRows"] + lobby_audit["invalidRows"]
    extreme_ratios = sum(1 for item in live_records if item["ratio"] > 1)
    low_volume = sum(1 for item in live_records if item["lobby"] < 100)
    pending_weeks = [item["week"] for item in synchronization if item["status"] == "pending"]
    quality = {
        "status": "ok" if source_invalid == 0 and zero_denominator == 0 else "review",
        "sourceDuplicatesConsolidated": source_duplicates,
        "invalidSourceRows": source_invalid,
        "zeroDenominatorRows": zero_denominator,
        "extremeRatiosFlagged": extreme_ratios,
        "lowVolumeRowsFlagged": low_volume,
        "latestCoverage": round(len(latest) / len(directory), 8) if directory else 0,
        "pendingWeeks": pending_weeks,
        "synchronization": synchronization,
    }

    # Cada nivel promedia el porcentaje calculado por tienda.
    weekly_rollups: dict[str, list[dict[str, Any]]] = {"region": [], "dm": []}
    for level in ("region", "dm"):
        grouped: defaultdict[tuple[int, str], list[dict[str, Any]]] = defaultdict(list)
        for item in live_records:
            grouped[(item["week"], item[level])].append(item)
        for (item_week, name), items in sorted(grouped.items()):
            rollup_fhw = sum(item["fhw"] for item in items)
            rollup_lobby = sum(item["lobby"] for item in items)
            weekly_rollups[level].append({
                "week": item_week,
                "name": name,
                "fhw": round(rollup_fhw, 6),
                "lobby": round(rollup_lobby, 6),
                "ratio": round(average_ratio(items), 8),
                "stores": len({item["ceco"] for item in items}),
            })

    example = next((item for item in live_records if item["ceco"] == "38101" and item["week"] == 30), None)
    audit = {
        "status": "ok",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "formula": "AVG(FHW / Bebidas Lobby) por tienda",
        "target": TARGET,
        "latestCompleteWeek": latest_week,
        "sources": {"fhw": fhw_audit, "lobby": lobby_audit, "reference": reference_audit},
        "joins": {
            "liveKeysFhw": len(fhw),
            "liveKeysLobby": len(lobby),
            "intersectedKeys": len(live_keys),
            "publishedLiveRows": len(live_records),
            "publishedHistoricalRows": len(historical_records),
            "excludedNoApplicableDirectory": excluded_no_directory,
            "zeroDenominator": zero_denominator,
        },
        "latest": {
            "stores": len(latest),
            "fhw": round(total_fhw, 6),
            "lobby": round(total_lobby, 6),
            "averageRatio": round(average_latest, 8),
            "storesAtTarget": sum(1 for item in latest if item["ratio"] > TARGET),
        },
        "quality": quality,
        "executiveWeeks": executive_weeks,
        "example38101w30": example,
    }
    payload = {
        "meta": {
            "title": "FHW · Cada Taza Cuenta",
            "version": "1.7.0",
            "generatedAt": audit["generatedAt"],
            "target": TARGET,
            "latestCompleteWeek": latest_week,
            "formula": audit["formula"],
            "weeks": sorted({item["week"] for item in all_records}),
            "months": [month for month in MONTHS if month in {item["month"] for item in all_records}],
            "monthWeeks": {
                month: sorted({item["week"] for item in all_records if item["month"] == month})
                for month in MONTHS if any(item["month"] == month for item in all_records)
            },
            "historyFiles": {},
            "weeksWithSynchronizedInputs": live_weeks,
            "weeksDirectOnly": sorted({item["week"] for item in historical_records}),
            "historyPolicy": "Semanas 1-29 usan el porcentaje directo por tienda; semanas 30+ calculan FHW/Lobby por tienda. Todos los alcances muestran el promedio de porcentajes válidos.",
            "updateState": {
                "readyWeeks": live_weeks,
                "pendingWeeks": pending_weeks,
                "latestReadyWeek": latest_week,
                "synchronization": synchronization,
            },
            "latestStores": len(latest),
            "organization": {
                "regions": len(hierarchy),
                "dms": len({item["dm"] for item in directory.values()}),
                "stores": len(directory),
                "hierarchy": hierarchy,
            },
            "eligibility": {
                "rule": reference_audit["applyRule"],
                "includedStores": len(directory),
                "excludedStores": reference_audit["excludedStores"],
            },
            "coverageByWeek": coverage_by_week,
            "executiveWeeks": executive_weeks,
            "weeklyRollups": weekly_rollups,
            "averageRollups": average_rollups,
            "quality": quality,
        },
        "records": sorted(live_records, key=lambda item: (item["week"], item["store"].casefold())),
        "historicalDm": [],
    }
    OUTPUT.mkdir(parents=True, exist_ok=True)
    history_output = OUTPUT / "history"
    history_next = OUTPUT / "history.next"
    history_backup = OUTPUT / "history.previous"
    if history_next.exists():
        shutil.rmtree(history_next)
    history_next.mkdir(parents=True)
    for month in payload["meta"]["months"]:
        month_records = [item for item in historical_records if item["month"] == month]
        if not month_records:
            continue
        weeks = {item["week"] for item in month_records}
        month_dm = [item for item in historical_dm if item["week"] in weeks]
        filename = f"{month.casefold()}.json"
        payload["meta"]["historyFiles"][month] = f"data/history/{filename}"
        atomic_json(history_next / filename, {"records": month_records, "historicalDm": month_dm})
    if history_backup.exists():
        shutil.rmtree(history_backup)
    if history_output.exists():
        history_output.replace(history_backup)
    history_next.replace(history_output)
    if history_backup.exists():
        shutil.rmtree(history_backup)
    atomic_json(OUTPUT / "fhw-dashboard.json", payload)
    atomic_json(OUTPUT / "data-audit.json", audit, compact=False)
    return audit


if __name__ == "__main__":
    print(json.dumps(build(), ensure_ascii=False))
